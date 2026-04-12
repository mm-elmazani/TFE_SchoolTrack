"""
Service d'import CSV/Excel pour les élèves.
Gère le parsing, la validation, la détection de doublons et l'insertion bulk.

Colonne optionnelle `classe` : si présente, l'élève est automatiquement assigné
à la classe correspondante (créée si elle n'existe pas encore).

Formats acceptés : CSV (.csv) et Excel (.xlsx).
"""

import csv
import io
import re
import unicodedata
import uuid
from typing import Optional, Tuple

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.school_class import ClassStudent, SchoolClass
from app.models.student import Student
from app.schemas.student import ImportError, StudentImportReport, StudentImportRow

# Colonnes acceptées (noms en français, insensibles à la casse et aux accents)
REQUIRED_COLUMNS = {"nom", "prenom"}
OPTIONAL_COLUMNS = {"email", "classe", "telephone"}
EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")

# Alias courants pour les colonnes (après suppression des accents et mise en minuscules)
COLUMN_ALIASES = {
    "prenom": "prenom",
    "prénom": "prenom",
    "firstname": "prenom",
    "first_name": "prenom",
    "nom": "nom",
    "lastname": "nom",
    "last_name": "nom",
    "mail": "email",
    "e-mail": "email",
    "email": "email",
    "classe": "classe",
    "class": "classe",
    "gsm": "telephone",
    "gsm eleve": "telephone",
    "gsm élève": "telephone",
    "telephone": "telephone",
    "téléphone": "telephone",
    "tel": "telephone",
    "phone": "telephone",
}


def _strip_accents(s: str) -> str:
    """Supprime les accents d'une chaîne."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def _normalize_header(raw: str) -> str:
    """Normalise un nom de colonne via la table d'alias (insensible casse/accents)."""
    cleaned = raw.strip().lower()
    if cleaned in COLUMN_ALIASES:
        return COLUMN_ALIASES[cleaned]
    stripped = _strip_accents(cleaned)
    if stripped in COLUMN_ALIASES:
        return COLUMN_ALIASES[stripped]
    return cleaned


def _detect_separator(sample: str) -> str:
    """Détecte le séparateur CSV (virgule ou point-virgule)."""
    if sample.count(";") >= sample.count(","):
        return ";"
    return ","


def _get_or_create_class(
    db: Session,
    class_name: str,
    school_id: Optional[uuid.UUID] = None,
) -> SchoolClass:
    """
    Retourne la classe portant ce nom dans l'école (insensible à la casse),
    ou la crée si elle n'existe pas encore.
    """
    query = select(SchoolClass).where(func.lower(SchoolClass.name) == class_name.lower())
    if school_id is not None:
        query = query.where(SchoolClass.school_id == school_id)
    existing = db.execute(query).scalar_one_or_none()

    if existing:
        return existing

    new_class = SchoolClass(name=class_name.strip(), school_id=school_id)
    db.add(new_class)
    db.flush()  # obtenir l'ID sans committer
    return new_class


def _validate_and_insert(
    rows: list[dict[str, str]],
    normalized_fields: set[str],
    db: Session,
    school_id: Optional[uuid.UUID] = None,
) -> StudentImportReport:
    """
    Logique commune de validation, détection de doublons et insertion bulk.
    `rows` : liste de dicts {colonne_normalisee: valeur} pour chaque ligne de données.
    """
    has_classe_column = "classe" in normalized_fields
    has_telephone_column = "telephone" in normalized_fields

    valid_rows: list[StudentImportRow] = []
    errors: list[ImportError] = []
    seen_in_file: set[Tuple[str, str]] = set()
    duplicates_in_file = 0

    for row_num, row in enumerate(rows, start=2):  # ligne 1 = header
        raw_last = row.get("nom", "").strip()
        raw_first = row.get("prenom", "").strip()
        raw_email = row.get("email", "").strip() if "email" in normalized_fields else ""
        raw_phone = row.get("telephone", "").strip() if has_telephone_column else ""
        raw_classe = row.get("classe", "").strip() if has_classe_column else ""

        if not raw_last and not raw_first:
            continue

        if not raw_last or not raw_first:
            errors.append(ImportError(
                row=row_num,
                content=f"{raw_last}, {raw_first}",
                reason="Nom ou prénom manquant"
            ))
            continue

        if raw_email and not EMAIL_REGEX.match(raw_email):
            errors.append(ImportError(
                row=row_num,
                content=f"{raw_last}, {raw_first}, {raw_email}",
                reason=f"Format email invalide : {raw_email}"
            ))
            continue

        key = (raw_last.lower(), raw_first.lower())
        if key in seen_in_file:
            duplicates_in_file += 1
            errors.append(ImportError(
                row=row_num,
                content=f"{raw_last}, {raw_first}",
                reason="Doublon dans le fichier"
            ))
            continue
        seen_in_file.add(key)

        valid_rows.append(StudentImportRow(
            last_name=raw_last,
            first_name=raw_first,
            email=raw_email or None,
            phone=raw_phone or None,
            classe=raw_classe or None,
        ))

    total_rows = len(rows)

    if not valid_rows:
        return StudentImportReport(
            total_rows=total_rows,
            inserted=0,
            rejected=len(errors),
            duplicates_in_file=duplicates_in_file,
            duplicates_in_db=0,
            errors=errors
        )

    # Detection doublons contre la BDD (colonnes chiffrees → comparaison Python, scopé par école)
    dup_query = select(Student.last_name, Student.first_name)
    if school_id is not None:
        dup_query = dup_query.where(Student.school_id == school_id)
    all_students = db.execute(dup_query).fetchall()
    existing_set = {
        (row[0].lower(), row[1].lower())
        for row in all_students
        if row[0] and row[1]
    }

    to_insert: list[StudentImportRow] = []
    duplicates_in_db = 0

    for student in valid_rows:
        key = (student.last_name.lower(), student.first_name.lower())
        if key in existing_set:
            duplicates_in_db += 1
            errors.append(ImportError(
                row=0,
                content=f"{student.last_name}, {student.first_name}",
                reason="Élève déjà présent en base de données"
            ))
        else:
            to_insert.append(student)

    if to_insert:
        db.add_all([
            Student(
                first_name=s.first_name,
                last_name=s.last_name,
                email=s.email,
                phone=s.phone,
                school_id=school_id,
            )
            for s in to_insert
        ])
        db.flush()

        if has_classe_column:
            _assign_classes(db, to_insert, school_id)

        db.commit()

    return StudentImportReport(
        total_rows=total_rows,
        inserted=len(to_insert),
        rejected=len(errors),
        duplicates_in_file=duplicates_in_file,
        duplicates_in_db=duplicates_in_db,
        errors=errors
    )


def parse_and_import_csv(
    content: bytes,
    db: Session,
    school_id: Optional[uuid.UUID] = None,
) -> StudentImportReport:
    """
    Parse un fichier CSV et importe les élèves.
    """
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1252")
    separator = _detect_separator(text.splitlines()[0] if text.splitlines() else "")

    reader = csv.DictReader(io.StringIO(text), delimiter=separator)

    if reader.fieldnames is None:
        return StudentImportReport(
            total_rows=0, inserted=0, rejected=0,
            duplicates_in_file=0, duplicates_in_db=0,
            errors=[ImportError(row=0, content="", reason="Fichier CSV vide ou illisible")]
        )

    field_map = {_normalize_header(f): f for f in reader.fieldnames}
    normalized_fields = set(field_map.keys())
    missing = REQUIRED_COLUMNS - normalized_fields
    if missing:
        return StudentImportReport(
            total_rows=0, inserted=0, rejected=0,
            duplicates_in_file=0, duplicates_in_db=0,
            errors=[ImportError(
                row=0, content=str(reader.fieldnames),
                reason=f"Colonnes manquantes : {', '.join(missing)}"
            )]
        )

    # Convertir chaque ligne en dict avec colonnes normalisées
    rows = []
    for raw_row in reader:
        rows.append({
            norm_col: (raw_row.get(orig_col, "") or "")
            for norm_col, orig_col in field_map.items()
        })

    return _validate_and_insert(rows, normalized_fields, db, school_id)


def parse_and_import_excel(
    content: bytes,
    db: Session,
    school_id: Optional[uuid.UUID] = None,
) -> StudentImportReport:
    """
    Parse un fichier Excel (.xlsx) et importe les élèves.
    Colonnes reconnues via la table d'alias (insensible casse/accents).
    Les colonnes non reconnues (badge, chambre, etc.) sont ignorées.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # Lire la première ligne comme en-tête
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        return StudentImportReport(
            total_rows=0, inserted=0, rejected=0,
            duplicates_in_file=0, duplicates_in_db=0,
            errors=[ImportError(row=0, content="", reason="Fichier Excel vide ou illisible")]
        )

    # Mapper index de colonne → nom normalisé (seules les colonnes reconnues)
    col_mapping: list[Tuple[int, str]] = []
    for idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        normalized = _normalize_header(str(cell_value))
        known = REQUIRED_COLUMNS | OPTIONAL_COLUMNS
        if normalized in known:
            col_mapping.append((idx, normalized))

    normalized_fields = {name for _, name in col_mapping}
    missing = REQUIRED_COLUMNS - normalized_fields
    if missing:
        wb.close()
        return StudentImportReport(
            total_rows=0, inserted=0, rejected=0,
            duplicates_in_file=0, duplicates_in_db=0,
            errors=[ImportError(
                row=0, content=str([h for h in header_row if h]),
                reason=f"Colonnes manquantes : {', '.join(missing)}"
            )]
        )

    # Lire les lignes de données
    rows = []
    for data_row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for idx, norm_name in col_mapping:
            val = data_row[idx] if idx < len(data_row) else None
            row_dict[norm_name] = str(val).strip() if val is not None else ""
        rows.append(row_dict)

    wb.close()
    return _validate_and_insert(rows, normalized_fields, db, school_id)


def _assign_classes(
    db: Session,
    students: list[StudentImportRow],
    school_id: Optional[uuid.UUID] = None,
) -> None:
    """
    Assigne chaque élève à sa classe après insertion.
    Crée la classe si elle n'existe pas encore.
    Ignore les élèves sans classe renseignée.
    """
    # Regrouper les élèves par classe pour minimiser les requêtes
    classes_map: dict[str, Optional[SchoolClass]] = {}

    for student in students:
        if not student.classe:
            continue

        class_name = student.classe.strip()
        if class_name not in classes_map:
            classes_map[class_name] = _get_or_create_class(db, class_name, school_id)

    if not classes_map:
        return

    # Recuperer les IDs des eleves (colonnes chiffrees → filtrage Python)
    target_keys = {
        (s.last_name.lower(), s.first_name.lower())
        for s in students if s.classe
    }
    all_rows = db.execute(
        select(Student.id, Student.last_name, Student.first_name)
    ).fetchall()
    student_id_map = {
        (row[1].lower(), row[2].lower()): row[0]
        for row in all_rows
        if row[1] and row[2] and (row[1].lower(), row[2].lower()) in target_keys
    }

    # Récupérer les assignations existantes pour éviter les doublons
    all_class_ids = [c.id for c in classes_map.values() if c]
    existing_assignments = set()
    if all_class_ids:
        existing_rows = db.execute(
            select(ClassStudent.class_id, ClassStudent.student_id)
            .where(ClassStudent.class_id.in_(all_class_ids))
        ).fetchall()
        existing_assignments = {(row[0], row[1]) for row in existing_rows}

    # Insérer les assignations manquantes
    assignments_to_insert = []
    for student in students:
        if not student.classe:
            continue

        school_class = classes_map.get(student.classe.strip())
        if not school_class:
            continue

        student_id = student_id_map.get(
            (student.last_name.lower(), student.first_name.lower())
        )
        if not student_id:
            continue

        if (school_class.id, student_id) not in existing_assignments:
            assignments_to_insert.append({
                "class_id": school_class.id,
                "student_id": student_id,
            })

    if assignments_to_insert:
        db.bulk_insert_mappings(ClassStudent, assignments_to_insert)
